import openpyxl
import csv
import typer

from typing_extensions import Annotated
from rich.progress import track
from pathlib import Path

def get_last_column(worksheet: openpyxl.worksheet) -> int:
    """
    Gets the last column containing a non-blank cell in row 2.

    :param worksheet: An openpyxl worksheet to search

    :return: an integer representing the last column with a non-blank cell in row 2 
    """

    last = 0

    for column in worksheet.iter_cols(min_row=2, max_row=2, values_only=True):
        for cell in column:
            if not cell:
                return last
        
        last = last + 1

    return last

def get_last_row(worksheet: openpyxl.worksheet) -> int:
    """
    Gets the last row containing a non-blank cell in column 1.

    :param worksheet: An openpyxl worksheet to search

    :return: an integer representing the last row with a non-blank cell in column 1 
    """

    last = 1

    for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
        for cell in row:
            if not cell:
                return last
            
            last = last + 1

    return last


def get_file_root(filename: str) -> str:
    """
    Gets the root name of the given filename with the path and extension removed.

    :param filename: A filename with optional leading path and suffix extension

    :return: string representing the name of the file minues the leading path and suffix extension
    """

    filepath = Path(filename)
    return filepath.stem

def main(filename: Annotated[str, typer.Argument(help="Excel filename of Ship Breaking Platform data")]) -> None:
    """
    Given a Ship Breaking Platform Excel-formatted spreadsheet, creates a new .csv of this
    data in the current working directory. 
    
    Quoting can be changed by altering the 'quoting' parameter in csv.writer().
    Leading and trailing whitespace is stripped from string data.

    :param filename: A filename of the Ship Breaking platform Excel spreadsheet
    """

    workbook = openpyxl.load_workbook(filename, read_only=False, data_only=True)
    worksheet = workbook.active

    rootname = get_file_root(filename)
    csvname = Path(rootname).with_suffix(".csv")

    last_data_column = get_last_column(worksheet)
    last_data_row = get_last_row(worksheet)

    with open(csvname, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_STRINGS)
        count = 0
    
        for row in track(worksheet.iter_rows(min_row=2, max_row=last_data_row, max_col=last_data_column, values_only=True), description=f"Writing '{csvname}'..."):
            cleaned_row = [
                cell.strip() if isinstance(cell, str) else cell
                for cell in row
            ]
            writer.writerow(cleaned_row)
            count = count +1

        print(f"Wrote {count} rows to '{csvname}'")

if __name__ == "__main__":
    typer.run(main)